from sklearn.linear_model import LinearRegression
import numpy as np
import pandas as pd
from dataclasses import dataclass
from datetime import datetime, timedelta
from openpyxl import load_workbook
import glob
import os
import matplotlib.pyplot as plt



A_max = 1
A_min = 0

# Angle array
theta_deg = np.linspace(0, 360, 1000)
time = theta_deg / 15  # convert angle → hours

# === Worst Day Eclipse (GEO) ===
eclipse_duration_h = 1.18        # 70.8 minutes
eclipse_center_h = 12.0          # midnight = GEO eclipse center

eclipse_half_angle = (eclipse_duration_h / 24 * 360) / 2
theta_eclipse_start = 180 - eclipse_half_angle
theta_eclipse_end   = 180 + eclipse_half_angle

# Convert eclipse window to time
t_eclipse_start = theta_eclipse_start / 15
t_eclipse_end   = theta_eclipse_end / 15

# === Create Sunlight Mask (1 in sun, 0 in eclipse) ===
sunlight_mask = np.where(
    (theta_deg >= theta_eclipse_start) & (theta_deg <= theta_eclipse_end),
    0,
    1
)

# === Area function with eclipse applied ===
def area(theta_deg):
    theta_rad = np.deg2rad(theta_deg)
    A = (A_max - A_min) * np.abs(np.cos(theta_rad)) + A_min
    return A * sunlight_mask    # <-- zero area during eclipse


# CONFIGURE MODEL

@dataclass
class ModelConfig:
    start_year: int = 2025
    years: int = 15
    eclipse_halfwidth_days: float = 22.5
    out_prefix: str = "solar_eff"
    eclipse_peak_minutes: float = 70.0
    decl_amp_deg: float = 23.44
    decl_phase_shift_day: int = 80
    ecc_amp: float = 0.033
    ecc_perihelion_shift_day: int = 3
    optical_drop_year1: float = 0.07
    cell_drop_year1: float = 0.03
    cell_drop_each_later_year: float = 0.02
    
     
# Helper functions
def to_doy(dates):
    dates = pd.to_datetime(dates)

    # If we got a scalar datetime → return a scalar int
    if isinstance(dates, pd.Timestamp):
        return dates.dayofyear

    # Otherwise it's an array → return array of ints
    return dates.dayofyear.values


def solar_distance_eff(n, amp=0.033, shift=3):
    n = np.asarray(n, dtype=float)
    return 1.0 + amp * np.cos(2 * np.pi * (n - shift) / 365.0)


def solar_angle_eff(n, decl_amp_deg=23.44, phase_shift=80):
    n = np.asarray(n, dtype=float)
    delta_deg = decl_amp_deg * np.sin(2 * np.pi * (n - phase_shift) / 365.0)
    return np.cos(np.deg2rad(delta_deg))


def eclipse_eff(n, centers=(80, 263), halfwidth=22.5, max_minutes=70.0):
    n = np.asarray(n, dtype=float)
    eclipse_minutes = np.zeros_like(n, dtype=float)
    for c in centers:
        d = np.abs(((n - c + 182.5) % 365) - 182.5)
        w = np.where(d <= halfwidth, 0.5 * (1 + np.cos(np.pi * d / halfwidth)), 0.0)
        eclipse_minutes += max_minutes * w
    return 1.0 - eclipse_minutes / (24.0 * 60.0)


def degradation_eff(day_index,
                    optical_drop_year1=0.07,
                    cell_drop_year1=0.03,
                    cell_drop_each_later_year=0.02):
    day_index = np.asarray(day_index, dtype=int)
    year_num = day_index // 365
    optical = np.where(year_num == 0, 1.0, 1.0 - optical_drop_year1)
    cells = (1.0 - cell_drop_year1) * np.power((1.0 - cell_drop_each_later_year), year_num)
    return optical * cells


# Core model
def build_timeseries(cfg: ModelConfig) -> pd.DataFrame:
    start_date = datetime(cfg.start_year, 1, 1)
    total_days = cfg.years * 365

    dates = [start_date + timedelta(days=i) for i in range(total_days)]
    doy = np.array([((to_doy(d) - 1) % 365) + 1 for d in dates])
    day_idx = np.arange(total_days)

    dist = solar_distance_eff(doy, cfg.ecc_amp, cfg.ecc_perihelion_shift_day)
    angle = solar_angle_eff(doy, cfg.decl_amp_deg, cfg.decl_phase_shift_day)
    ecl = eclipse_eff(doy, (80, 263), cfg.eclipse_halfwidth_days, cfg.eclipse_peak_minutes)
    deg = degradation_eff(day_idx, cfg.optical_drop_year1,
                          cfg.cell_drop_year1, cfg.cell_drop_each_later_year)

    total = dist * angle * ecl * deg
    total = np.clip(total, 0.0, 1.2)

    df = pd.DataFrame({
        "date": dates,
        #"day_of_year": doy,
        "distance_eff": dist,
        "angle_eff": angle,
        "eclipse_eff": ecl,
        "degradation_eff": deg,
        "total_eff": total
    }).set_index("date")

    return df


def annual_summary(df: pd.DataFrame) -> pd.DataFrame:
    out = df["total_eff"].resample("Y").agg(["min", "mean", "max"])
    
    out.index = np.arange(1, len(out) + 1)

    return out



# Plot: Total with regressed-degradation smoothing
def plot_total_with_regressed_degradation(df: pd.DataFrame, cfg: ModelConfig) -> str:
    """
    Create a 'smoothed' total efficiency curve by replacing the degradation factor
    with its 10-year linear-regression prediction, and plot both the raw daily
    total_eff and the smoothed (regression-based) total.
    """
    path = f"{cfg.out_prefix}_total_with_regressed_degradation.png"

    # --- Fit linear regression on the full 10-year degradation (t in years) ---
    N = len(df)
    t_years = np.linspace(0.0, float(cfg.years), N).reshape(-1, 1)  # shape (N,1)
    deg_daily = df["degradation_eff"].values.reshape(-1, 1)

    model = LinearRegression().fit(t_years, deg_daily)
    deg_pred = model.predict(t_years).flatten()

    slope = float(model.coef_[0])
    intercept = float(model.intercept_)

    # --- Build regressed total_eff (using predicted degradation) ---
    dist = df["distance_eff"].values
    angle = df["angle_eff"].values
    ecl = df["eclipse_eff"].values

    total_regressed = dist * angle * ecl * deg_pred
    total_regressed = np.clip(total_regressed, 0.0, 1.2)


# EXCEL LOADING

def load_excel_values(path):
    df = pd.read_excel(path, sheet_name="Analysis", header=None)
    idx = df[df[0] == "BEST COATING"].index[0]
    global energy_worst_day
    energy_worst_day = float(df.iloc[idx + 13, 1])
    global heater_power
    heater_power = float(df.iloc[idx + 4, 1])
        
    return {
        "name": df.iloc[idx + 1, 1],
        "absorptivity": float(df.iloc[idx + 2, 1]),
        "emissivity": float(df.iloc[idx + 3, 1]),
        "heater_power": float(df.iloc[idx + 4, 1]),
        "annual_kwh": float(df.iloc[idx + 5, 1]),
        "energy_worst_day": float(df.iloc[idx + 13, 1])
    }


# BUILD TIMESERIES WITH INTERPOLATION + REGRESSION

def build_efficiency(cfg):
    total_days = cfg.years * 365

    dates = np.array([datetime(cfg.start_year, 1, 1) + timedelta(days=i)
                      for i in range(total_days)])

    doy = to_doy(dates)
    idx = np.arange(total_days)

    dist = solar_distance_eff(doy, cfg.ecc_amp, cfg.ecc_perihelion_shift_day)
    angle = solar_angle_eff(doy, cfg.decl_amp_deg, cfg.decl_phase_shift_day)
    ecl = eclipse_eff(doy, [80, 263], cfg.eclipse_halfwidth_days, cfg.eclipse_peak_minutes)
    deg = degradation_eff(idx, cfg.optical_drop_year1, cfg.cell_drop_year1, cfg.cell_drop_each_later_year)

    total_eff = np.clip(dist * angle * ecl * deg, 0, 1.2)

    df = pd.DataFrame({"date": dates, "total_eff": total_eff}).set_index("date")

    
    # 1) LINEAR INTERPOLATION
    
    # (Your data is already daily, but interpolation makes the model robust.)
    df["total_eff_interp"] = df["total_eff"].interpolate(method="linear")

    
    # 2) LINEAR REGRESSION (efficiency vs mission day index)
    
    slope, intercept = np.polyfit(idx, total_eff, 1)
    trend = slope * idx + intercept

    df["eff_trendline"] = trend

    return df




# -----PANEL AREA (USING USABLE POWER = +10%) ----
internal_energy_use = 10000 # W

I_sun = 1361.0               # W/m²
absobtivity = 0.86            # solar cell absorptivity
eta_panel = 0.40             # solar cell efficiency
eta_sys = 0.679               # system efficiency
eta_total = eta_panel * eta_sys   # effective conversion efficiency

def surface_area(energy_heater):
    global energy_J
    energy_J =  (energy_heater) + (internal_energy_use * 86400)  # J needed per day
    power_per_m2 = I_sun * absobtivity * eta_total * area(theta_deg)   

    # integrate over full day
    dt = 86400 / len(power_per_m2)
    energy_per_m2_day = np.sum(power_per_m2) * dt     # J per m² per day
    global area_required 
    area_required = energy_J / energy_per_m2_day

    # 10% margin
    global area_required_10
    area_required_10 = area_required * 1.10
    
    print(energy_heater)

    return area_required_10


# POWER TIMESERIES

def power_timeseries(df_eff, area, absorbed_per_m2):
    df_eff["power_generated"] = area * absorbed_per_m2 * df_eff["total_eff"]
    return df_eff


# MAIN

def main():
    parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_pattern = os.path.join(parent_dir, "satellite_analysis*.xlsx")
    excel_files = glob.glob(excel_pattern)

    if not excel_files:
        raise FileNotFoundError(f"No satellite_analysis*.xlsx files found in {parent_dir}")

    excel_path = max(excel_files, key=os.path.getctime)
    vals = load_excel_values(excel_path)

    cfg = ModelConfig()
    df_eff = build_efficiency(cfg)

    # Calculate worst efficiency and year
    worst_eff = df_eff["total_eff"].min()
    worst_idx = df_eff["total_eff"].idxmin()
    
    # Calculate panel area from energy requirements
    area = surface_area(vals["energy_worst_day"])
    
    # Calculate usable power (area * solar intensity * efficiency * absorptivity)
    usable_power = area * I_sun * absobtivity * eta_total

    return df_eff, area, worst_eff, usable_power, excel_path


if __name__ == "__main__":
    df_out, panel_area, worst_eff, usable_power, excel_path = main()

    wb = load_workbook(excel_path)
    ws = wb["Analysis"]

    write_row = 44  # fixed row

    ws[f"A{write_row}"] = "Worst Efficiency"
    ws[f"B{write_row}"] = float(worst_eff)

    ws[f"A{write_row+2}"] = "Required Panel Area (m²)"
    ws[f"B{write_row+2}"] = float(area_required_10)

    ws[f"A{write_row+3}"] = "Usable Power (W)"
    ws[f"B{write_row+3}"] = float(usable_power)

    wb.save(excel_path)

    print(f"Results written successfully at rows {write_row}–{write_row+3}.")


energy_J =  (energy_worst_day) + (internal_energy_use * 86400)

percentage = ((energy_J / 86400) / (I_sun * eta_total * absobtivity * area_required_10))

A_values = area(theta_deg)
mask = A_values < percentage * A_max
indices = np.where(mask)[0]

intervals = []
if len(indices) > 0:
    start = indices[0]
    for i in range(1, len(indices)):
        if indices[i] != indices[i - 1] + 1:
            end = indices[i - 1]
            intervals.append((time[start], time[end]))
            start = indices[i]
    intervals.append((time[start], time[indices[-1]]))

print("Intervals when battery is needed:")
for (t1, t2) in intervals:
    print(f"From {t1:.2f} h to {t2:.2f} h")


plt.figure(figsize=(10, 5))
plt.plot(time, area(theta_deg), lw=2, label="Visible Solar Area")

# Shade eclipse
plt.axvspan(t_eclipse_start, t_eclipse_end, color='gray', alpha=0.3,
            label="Eclipse (Area=0)")

plt.title('Visible Solar Panel Area vs Time')
plt.xlabel('Time, t (Hours)')
plt.ylabel('Visible surface area, A (%)')
plt.axhline(percentage, color='r', linestyle='--', label="Required Power")
plt.xlim(0, 24)
plt.ylim(0, A_max * 1.05)
plt.grid(alpha=0.3)
plt.legend()
plt.show()

# ----END OF PANEL AREA CALCULATION ----


# ---- START OF EFFICIENCY AND DEGREDATION MODEL -----
# Plot: Total with regressed-degradation smoothing
def plot_total_with_regressed_degradation(df: pd.DataFrame, cfg: ModelConfig) -> str:
    """
    Create a 'smoothed' total efficiency curve by replacing the degradation factor
    with its 10-year linear-regression prediction, and plot both the raw daily
    total_eff and the smoothed (regression-based) total.
    """
    path = f"{cfg.out_prefix}_total_with_regressed_degradation.png"

    # --- Fit linear regression on the full 10-year degradation (t in years) ---
    N = len(df)
    t_years = np.linspace(0.0, float(cfg.years), N).reshape(-1, 1)  # shape (N,1)
    deg_daily = df["degradation_eff"].values.reshape(-1, 1)

    model = LinearRegression().fit(t_years, deg_daily)
    deg_pred = model.predict(t_years).flatten()

    slope = float(model.coef_[0])
    intercept = float(model.intercept_)

    # --- Build regressed total_eff (using predicted degradation) ---
    dist = df["distance_eff"].values
    angle = df["angle_eff"].values
    ecl = df["eclipse_eff"].values

    total_regressed = dist * angle * ecl * deg_pred
    total_regressed = np.clip(total_regressed, 0.0, 1.2)

    # --- Plot original daily total and regressed (smoothed) total ---
    plt.figure(figsize=(13, 5))
    # raw daily total (light)
    t = np.linspace(1, cfg.years, len(df))
    plt.plot(t, df["total_eff"].values, label="Daily total_eff (raw)", color="lightgrey", alpha=0.7, linewidth=0.8)
    # smoothed total from regressed degradation (strong)
    plt.plot(t, total_regressed, label="Total (smoothed via regressed degradation)", color="tab:blue", linewidth=2.2)

    plt.title(f"Total Efficiency — Raw daily vs Smoothed (regression-based degradation) — {cfg.years} years")
    plt.xlabel("Year")
    plt.ylabel("Efficiency (fraction of initial)")
    plt.legend()
    plt.grid(True, linestyle="--", alpha=0.3)

    # annotate degradation regression equation
    xy_loc = (df.index[int(N * 0.02)], max(total_regressed) * 0.98)
    plt.text(xy_loc[0], xy_loc[1],
             f"Degradation fit (10y):\n  deg = {slope:.6f}·t + {intercept:.6f}\n(t in years)",
             color="black", fontsize=9, bbox=dict(facecolor="white", alpha=0.8, edgecolor="none"))

    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.show()

    return path


# (Optional) plot degradation + regression
def plot_degradation_10yr_with_regression(df: pd.DataFrame, cfg: ModelConfig) -> str:
    path = f"{cfg.out_prefix}_10yr_degradation_regression.png"
    N = len(df)
    t_years = np.linspace(0.0, float(cfg.years), N)
    deg_daily = df["degradation_eff"].values

    model = LinearRegression().fit(t_years.reshape(-1, 1), deg_daily)
    deg_pred = model.predict(t_years.reshape(-1, 1))

    slope = float(model.coef_[0])
    intercept = float(model.intercept_)

    plt.figure(figsize=(12, 5))
    t = np.linspace(1, cfg.years, len(df))
    plt.plot(t, deg_daily, label="Daily degradation", color="red", linewidth=0.9, alpha=0.8)
    plt.plot(t, deg_pred, "k--", linewidth=2.0, label=f"Linear fit: deg = {slope:.6f}·t + {intercept:.6f}")
    plt.title("15-Year Degradation (daily) with Linear Regression")
    plt.xlabel("Year")
    plt.ylabel("Degradation efficiency")
    plt.grid(True, linestyle="--", alpha=0.3)
    plt.legend()
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.show()
    return path


# MAIN
def main():
    cfg = ModelConfig()
    df = build_timeseries(cfg)

    # Save CSVs
    csv_daily = f"{cfg.out_prefix}_daily_{cfg.start_year}_{cfg.start_year + cfg.years - 1}.csv"
    df.to_csv(csv_daily)
    global annual
    annual = annual_summary(df)
    annual.to_csv(f"{cfg.out_prefix}_annual_summary.csv")

    print(f"[OK] Saved daily CSV -> {csv_daily}")
    print(f"[OK] Saved annual summary -> {cfg.out_prefix}_annual_summary.csv")
    print("\nAnnual min/mean/max:")
    print(annual.round(4))

    # Plots
    global p1, p2
    p1 = plot_total_with_regressed_degradation(df, cfg)
    p2 = plot_degradation_10yr_with_regression(df, cfg)  # diagnostic (optional)

    print(f"[OK] Saved plots -> {p1}, {p2}")

    # Summary
    print("\nEfficiency range:",
          f"{df['total_eff'].min():.3f} – {df['total_eff'].max():.3f}")
    print(f"Mean lifetime efficiency ({cfg.years}y) ≈ {df['total_eff'].mean():.3f}")


# ---- END OF EFFICIENCY AND DEGREDATION MODEL
if __name__ == "__main__":
    main()
    
from openpyxl.drawing.image import Image

def export_to_excel(excel_path, annual_df, plot_paths):
    wb = load_workbook(excel_path)
    ws = wb["Analysis"]

    # Write annual efficiency table
    start_row = 60
    ws[f"A{start_row}"] = "Year"
    ws[f"B{start_row}"] = "Min"
    ws[f"C{start_row}"] = "Mean"
    ws[f"D{start_row}"] = "Max"

    for i, (yr, row) in enumerate(annual_df.iterrows(), start=start_row+1):
        ws[f"A{i}"] = yr
        ws[f"B{i}"] = float(row["min"])
        ws[f"C{i}"] = float(row["mean"])
        ws[f"D{i}"] = float(row["max"])

    # Insert images
    img_row = start_row + len(annual_df) + 3
    for p in plot_paths:
        img = Image(p)
        ws.add_image(img, f"F{img_row}")
        img_row += 30

    wb.save(excel_path)
    
    
export_to_excel(excel_path, annual, [p1, p2])
    

