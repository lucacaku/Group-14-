#import any necessary libraries
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import pandas as pd
from datetime import datetime
import io

#define constants and parameters
sigma = 5.67e-8
solar_constant = 1361
albedo = 0.3
Gravitational_constant = 6.67430e-11
mass_of_earth = 5.972e24

threshold_low = 293
threshold_high = 313
average_threshold = (threshold_low + threshold_high) / 2.0
c = 900.0  #specific heat capacity J/(kg·K)

#user defined satellite geometry and mass plus error handling to prevent unrealistic values
def _get_float_input(prompt, min_val, max_val):
    while True:
        try:
            s = input(f"{prompt} [{min_val} - {max_val}]: ")
            val = float(s)
        except ValueError:
            print("Invalid input — enter a numeric value.")
            continue
        if not (min_val <= val <= max_val):
            print(f"Out of range — enter a value between {min_val} and {max_val}.")
            continue
        return val

height = _get_float_input("Input a height for your satellite (m)", 1.0, 5.0)
width  = _get_float_input("Input a width for your satellite (m)", 1.0, 5.0)
depth  = _get_float_input("Input a depth for your satellite (m)", 3.0, 7.0)
mass   = _get_float_input("Input a mass for your satellite (kg)", 1000.0, 4000.0)

print("Our code takes a moment to run, bear with us!")

#initialise coating properties array
coatings = {
    'Polymide Nanofiber Films': {'alpha': 0.004, 'epsilon': 0.93},
    'Ta2O5, SiN, SiO2 Film': {'alpha': 0.110, 'epsilon': 0.750},
    'Hughson White Paint A276': {'alpha': 0.26, 'epsilon': 0.88},
    'Bare Polished Aluminum': {'alpha': 0.4, 'epsilon': 0.04}, #establish baseline coating
}

#define orbital parameters and time arrays
altitude = 35786e3
earth_radius = 6371e3
year_seconds = 365.25 * 24 * 3600

orbital_radius = earth_radius + altitude
orbital_period = 2 * np.pi * np.sqrt(orbital_radius**3 / (Gravitational_constant * mass_of_earth)) #Keplers 3rd law (~23.93 hours)

dt = 10 * 60 #small timestep so accurate, not too small to be slow
day_start = 0
sim_days = 365
simulation_end = (day_start + sim_days) * 24 * 3600

time_orbit = np.linspace(0, orbital_period, 1000)
time_year = np.linspace(0, year_seconds, int(year_seconds / dt))

time_sweep_start = simulation_start
time_sweep_days = 3
time_sweep = np.arange(time_sweep_start, time_sweep_start + time_sweep_days * 24 * 3600 + dt, dt)

time_long_start = day_start

#eclipse and area functions
def eclipse_mask(t):
    day = np.atleast_1d(t / 86400) % 365.25
    season1 = (day > 59) & (day < 99) #defined eclipse seasons sync with earth equinoxes due to GEO orbit
    season2 = (day > 239) & (day < 279)
    in_season = season1 | season2

    eclipse_duration = 0.0708 * 86400
    orbit_phase = np.atleast_1d(t) % orbital_period #handles scalar and array inputs for time
    in_eclipse = orbit_phase < eclipse_duration

    return in_season & in_eclipse

def set_interps_for(time_array):
    global A_pres_interp, sunlight_interp, A_earth_interp

    sunlight = np.where(eclipse_mask(time_array), 0, 1) #binary, 0 if shade, 1 if sun

    #tile presented area over orbits to match time_array length
    theta_rad = np.pi/4 * np.sin(2 * np.pi * time_orbit / orbital_period)
    A_presented_orbit = height * (np.abs(width * np.cos(theta_rad)) + np.abs(depth * np.sin(theta_rad))) #formula assumes single plane rotation

    reps = int(np.ceil(len(time_array) / len(A_presented_orbit)))
    A_presented = np.tile(A_presented_orbit, reps)[:len(time_array)] 

    half_angle_earth = np.arcsin(earth_radius / orbital_radius)
    A_earth_orbit = np.pi * (earth_radius**2) * (1 - np.cos(half_angle_earth))
    A_earth = A_earth_orbit * sunlight

    A_pres_interp = interp1d(time_array, A_presented, kind='linear', fill_value='extrapolate')
    sunlight_interp = interp1d(time_array, sunlight, kind='nearest', fill_value='extrapolate')
    A_earth_interp = interp1d(time_array, A_earth, kind='nearest', fill_value='extrapolate')

#thermal modelling functions
def thermal_rhs(t, T, alpha, epsilon, heater_state, heater_power):
    T = np.clip(float(T), 0, 750) #prevent unphysical temperatures

    A_pres = float(A_pres_interp(t))
    is_sunlit = float(sunlight_interp(t))
    A_earth_local = float(A_earth_interp(t))

    Q_solar = alpha * solar_constant * is_sunlit * A_pres
    Q_albedo = solar_constant * albedo * (A_earth_local / (4.0 * np.pi * orbital_radius**2)) #assumes reflected sun is uniform hemisphere
    Q_ir_loss = epsilon * sigma * (T**4) * A_pres

    if T < average_threshold: #heater turns on at low threshold, off at high threshold
        heater_state = 1
    elif T >= average_threshold:
        heater_state = 0 #not threshold_high - delta, prevents energy waste by rapid switching

    Q_heater = heater_power * heater_state
    Q_net = Q_solar + Q_albedo + Q_heater - Q_ir_loss
    dTdt = Q_net / (mass * c)

    return dTdt, heater_state, Q_heater

def heat_balance_RK4_with_heater(alpha, epsilon, time_array, heater_power, T_init=average_threshold): #runge-kutta 4th order
    n = len(time_array)
    T = np.zeros(n)
    heater_state = np.zeros(n)
    heater_energy = np.zeros(n)

    T[0] = T_init
    heater_state[0] = 0

    for i in range(n - 1):
        t = time_array[i]
        h = time_array[i + 1] - time_array[i]

        #rk4 slopes
        k1, h1_state, Qh1 = thermal_rhs(t, T[i], alpha, epsilon, heater_state[i], heater_power) 
        k2, h2_state, Qh2 = thermal_rhs(t + h/2, T[i] + h/2*k1, alpha, epsilon, h1_state, heater_power)
        k3, h3_state, Qh3 = thermal_rhs(t + h/2, T[i] + h/2*k2, alpha, epsilon, h2_state, heater_power)
        k4, h4_state, Qh4 = thermal_rhs(t + h, T[i] + h*k3, alpha, epsilon, h3_state, heater_power)

        slope = (k1 + 2*k2 + 2*k3 + k4) / 6.0
        T[i+1] = T[i] + h * slope
        heater_state[i+1] = h4_state
        heater_energy[i+1] = heater_energy[i] + Qh4 * h

    total_energy_kWh = heater_energy[-1] / 3.6e6 #convert from joules to kWh
    return T, heater_state, total_energy_kWh

set_interps_for(time_sweep)

#heater power sweep to find optimal power for each coating
heater_powers = np.arange(0000, 15000, 200) # tests 0-15 kW in 200 W increments
results = {}


for name, props in coatings.items():
    alpha = props['alpha']
    epsilon = props['epsilon']

    # initialize container for this coating's results
    results[name] = {}
    energy_vs_power = []  # store total heater energy for each power level

    # print(f"\n=== {name} ===")
    for power in heater_powers:
        def thermal_rhs_power(t, T, alpha=alpha, epsilon=epsilon, heater_state=0):
            T = float(T)
            A_pres = float(A_pres_interp(t))
            is_sunlit = float(sunlight_interp(t))
            A_earth_local = float(A_earth_interp(t))

            # energy balance
            Q_solar = alpha * solar_constant * is_sunlit * A_pres
            Q_albedo = solar_constant * albedo * (A_earth_local / (4.0 * np.pi * orbital_radius**2))
            Q_ir_loss = epsilon * sigma * (T**4) * A_pres

            if T < average_threshold:
                heater_state = 1
            elif T >= average_threshold:
                heater_state = 0

            Q_heater = power * heater_state
            Q_net = Q_solar + Q_albedo + Q_heater - Q_ir_loss

            dTdt = Q_net / (mass * c)
            return dTdt, heater_state, Q_heater

        # integrate temperature over the short sweep time array
        n = len(time_sweep)
        T = np.zeros(n)
        heater_state = np.zeros(n)
        heater_energy = np.zeros(n)

        T[0] = average_threshold
        heater_state[0] = 0

        for i in range(n - 1):
            t = time_sweep[i]
            h = time_sweep[i + 1] - time_sweep[i]

            # RK4 integration
            k1, h1_state, Qh1 = thermal_rhs_power(t, T[i])
            k2, h2_state, Qh2 = thermal_rhs_power(t + h/2, T[i] + h/2*k1)
            k3, h3_state, Qh3 = thermal_rhs_power(t + h/2, T[i] + h/2*k2)
            k4, h4_state, Qh4 = thermal_rhs_power(t + h, T[i] + h*k3)

            slope = (k1 + 2*k2 + 2*k3 + k4) / 6.0
            T[i+1] = T[i] + h * slope
            heater_state[i+1] = h4_state
            heater_energy[i+1] = heater_energy[i] + Qh4 * h

        total_energy_kWh = heater_energy[-1] / 3.6e6
        energy_vs_power.append(total_energy_kWh)
    
    # After the inner loop over power
    results[name]["heater_powers"] = heater_powers
    results[name]["energy_vs_power"] = energy_vs_power

for name in results:
    hp = np.array(results[name]["heater_powers"])
    evp = np.array(results[name]["energy_vs_power"])

    if len(evp) < 2:
        continue

    slope = evp[1] - evp[0]

    mask = np.ones_like(evp, dtype=bool)

    for i in range(len(evp)):
        expected = slope * i

        if abs(evp[i] - expected) < 1e-6:
            mask[i] = False

    hp_clean = hp[mask]
    evp_clean = evp[mask]

    results[name]["heater_powers"] = hp_clean
    results[name]["energy_vs_power"] = evp_clean


for name in results:
    hp = results[name]["heater_powers"]
    evp = results[name]["energy_vs_power"]

    if len(hp) == 0:
        results[name]["optimal_power_W"] = 0
        continue

    idx = np.argmin(evp)
    optimal = float(hp[idx])

    results[name]["optimal_power_W"] = optimal
    print(f"{name} → optimal power = {optimal} W")

def get_optimal(name):
    entry = results.get(name)
    if entry is None:
        return 0
    return entry.get("optimal_power_W", 0)

min_power_1 = get_optimal("Polymide Nanofiber Films")
min_power_2 = get_optimal("Ta2O5, SiN, SiO2 Film")
min_power_3 = get_optimal("Hughson White Paint A276")
min_power_4 = get_optimal("Bare Polished Aluminum")

for name, data in results.items():
    plt.figure(figsize=(12,5))

    plt.plot(data["heater_powers"]/1000, data["energy_vs_power"], 'o-', label="Heater Energy")
    plt.xlabel("Heater Power (kW)")
    plt.ylabel("Total Energy Used (kWh)")
    plt.title(f"{name} – Heater Energy vs Power")
    plt.grid(True)
    plt.tight_layout()
    plt.legend()

plt.show()

print("Our code is still running!")


#is there any way to remove this? to just append to our old array?
coatings = {
    'Polymide Nanofiber Films': {'alpha': 0.004, 'epsilon': 0.93, 'power': min_power_1},
    'Ta2O5, SiN, SiO2 Film': {'alpha': 0.110, 'epsilon': 0.750, 'power': min_power_2},
    'Hughson White Paint A276': {'alpha': 0.26, 'epsilon': 0.88, 'power': min_power_3},
    'Bare Polished Aluminum': {'alpha': 0.4, 'epsilon': 0.04, 'power': min_power_4}
}

#full year simulation with optimal heater powers
set_interps_for(time_year)

for name, props in coatings.items():
    alpha, epsilon = props['alpha'], props['epsilon']
    power = props.get('power',0) #use previously found optimal power

    T_with, heater_state, energy_kWh = heat_balance_RK4_with_heater(alpha, epsilon, time_year, power)

    results[name].update({
        "T": T_with,
        "heater": heater_state,
        "energy_kWh": energy_kWh
    })

    print(f"{name}: {energy_kWh:.2f} kWh/year at {power:.0f} W Heater Power")

#storing results and analysis
in_band_fraction = {name: np.mean((data["T"] >= threshold_low) & (data["T"] <= threshold_high))
                    for name, data in results.items()}

best_name = max(in_band_fraction, key=in_band_fraction.get)
T_best = results[best_name]["T"]

#analysis of deviations from temperature band
below = T_best[T_best < threshold_low]
above = T_best[T_best > threshold_high]

avg_below = np.mean(threshold_low - below) if len(below) > 0 else 0.0
max_below = np.max(threshold_low - below) if len(below) > 0 else 0.0
avg_above = np.mean(above - threshold_high) if len(above) > 0 else 0.0
max_above = np.max(above - threshold_high) if len(above) > 0 else 0.0

print(f"\n--- Deviation analysis for {best_name} ---")
print(f"Time within band: {in_band_fraction[best_name]*100:.2f}%")
print(f"Below threshold: {len(below)/len(T_best)*100:.2f}%")
print(f"  Avg: {avg_below:.2f} K, Max: {max_below:.2f} K")
print(f"Above threshold: {len(above)/len(T_best)*100:.2f}%")
print(f"  Avg: {avg_above:.2f} K, Max: {max_above:.2f} K")

#scoring based off energy use, in-band time, heater power
scorable = [n for n in results.keys() if n != "Bare Polished Aluminum"]

powers = np.array([results[n]["optimal_power_W"] for n in scorable], dtype=float)
energies = np.array([results[n]["energy_kWh"] for n in scorable], dtype=float)
inbands = np.array([in_band_fraction[n] for n in scorable], dtype=float)

def norm_lower_better(x):
    xmin, xmax = x.min(), x.max()
    if np.isclose(xmax, xmin):
        return np.ones_like(x)
    return 1.0 - (x - xmin) / (xmax - xmin)

def norm_higher_better(x):
    xmin, xmax = x.min(), x.max()
    if np.isclose(xmax, xmin):
        return np.ones_like(x)
    return (x - xmin) / (xmax - xmin)

norm_p = norm_lower_better(powers)
norm_e = norm_lower_better(energies)
norm_i = norm_higher_better(inbands)

w_inband, w_energy, w_power = 0.3, 0.3, 0.4

scores = {}
for idx, name in enumerate(scorable):
    combined = (w_inband * norm_i[idx] + w_energy * norm_e[idx] + w_power * norm_p[idx])
    scores[name] = {
        "in_band": in_band_fraction[name] * 100,
        "energy_kWh": results[name]["energy_kWh"],
        "optimal_power_W": results[name]["optimal_power_W"],
        "combined_score": combined
    }
    print(f"{name}: {in_band_fraction[name]*100:.2f}% in-band, {results[name]['energy_kWh']:.2f} kWh/year, "
          f"{results[name]['optimal_power_W']:.0f} W heater → score={combined:.3f}")

best_balanced = max(scores, key=lambda x: scores[x]["combined_score"]) if scores else None
print(f"\n→ Best balanced choice: {best_balanced}")

#window around eclipse season for plotting
def plot_eclipse_season():
    plt.figure(figsize=(10, 5))
    start_day, end_day = 45, 75
    start_idx = np.searchsorted(time_year, start_day * 86400.0)
    end_idx = np.searchsorted(time_year, end_day * 86400.0)

    for name, data in results.items():
        plt.plot(time_year[start_idx:end_idx]/86400.0, data["T"][start_idx:end_idx], label=f"{name}")

    plt.fill_between([start_day, end_day], [threshold_low, threshold_low], [threshold_high, threshold_high], 
                    color='orange', alpha=0.3)
    plt.fill_betweenx([0,750], 60, 75, color = 'gray', alpha = 0.3)
    plt.xlabel("Time (days)")
    plt.ylabel("Temperature (K)")
    plt.title("Temperature Evolution – Eclipse Season")
    plt.legend(fontsize='small')
    plt.grid(True)
    return plt.gcf()

#six month comparison of worst and best coatings
def create_six_month_plot():
    plt.figure(figsize=(10, 6))
    months_6_idx = np.searchsorted(time_year, 182.625 * 86400.0)

    for name in [best_balanced, "Bare Polished Aluminum"]:
        plt.plot(2 * time_year[:months_6_idx] / year_seconds * 6.0, results[name]["T"][:months_6_idx],
                label=f"{name} ({in_band_fraction[name]*100:.1f}%)")

    plt.axhline(threshold_low, linestyle='--', color='gray')
    plt.axhline(threshold_high, linestyle='--', color='gray')
    plt.xlabel("Month")
    plt.xticks(ticks=np.arange(0, 6, 1), labels=['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'])
    plt.ylabel("Temperature (K)")
    plt.title("Six-Month Temperature Profile")
    plt.legend()
    plt.grid(True)
    return plt.gcf()

#create output data for excel
output_data = []

#satellite dimensions section
output_data.append(['SATELLITE DIMENSIONS', '', ''])
output_data.append(['Height', f"{height:.2f}", 'm'])
output_data.append(['Width', f"{width:.2f}", 'm'])
output_data.append(['Depth', f"{depth:.2f}", 'm'])
output_data.append(['Mass', f"{mass:.0f}", 'kg'])
output_data.append(['Specific Heat Capacity', f"{c:.1f}", 'J/kg·K'])
output_data.append(['', '', ''])

#orbital geometry section
output_data.append(['ORBITAL GEOMETRY', '', ''])
output_data.append(['Orbital Altitude', f"{altitude/1e3:.0f}", 'km'])
output_data.append(['Orbital Radius', f"{orbital_radius/1e3:.0f}", 'km'])
output_data.append(['Orbital Period', f"{orbital_period/3600:.2f}", 'hours'])
output_data.append(['Earth Radius', f"{earth_radius/1e3:.0f}", 'km'])
output_data.append(['', '', ''])

#thermal parameters section
output_data.append(['THERMAL PARAMETERS', '', ''])
output_data.append(['Temperature Threshold Low', f"{threshold_low:.0f}", 'K'])
output_data.append(['Temperature Threshold High', f"{threshold_high:.0f}", 'K'])
output_data.append(['Solar Constant', f"{solar_constant}", 'W/m²'])
output_data.append(['Albedo', f"{albedo}", ''])
output_data.append(['Stefan-Boltzmann Constant', f"{sigma}", 'W/m²·K⁴'])
output_data.append(['', '', ''])

#best coating summary section
output_data.append(['BEST COATING', '', ''])
output_data.append(['Coating Name', best_balanced, ''])
output_data.append(['Solar Absorptivity', f"{coatings[best_balanced]['alpha']:.4f}", ''])
output_data.append(['Thermal Emissivity', f"{coatings[best_balanced]['epsilon']:.4f}", ''])
output_data.append(['Optimal Heater Power', f"{results[best_balanced]['optimal_power_W']:.0f}", 'W'])
output_data.append(['Annual Energy Usage', f"{results[best_balanced]['energy_kWh']:.2f}", 'kWh'])
output_data.append(['Time In-Band', f"{in_band_fraction[best_balanced]*100:.2f}", '%'])
output_data.append(['Time Below Threshold', f"{len(below)/len(T_best)*100:.2f}", '%'])
output_data.append(['Avg Deviation Below', f"{avg_below:.2f}", 'K'])
output_data.append(['Max Deviation Below', f"{max_below:.2f}", 'K'])
output_data.append(['Time Above Threshold', f"{len(above)/len(T_best)*100:.2f}", '%'])
output_data.append(['Avg Deviation Above', f"{avg_above:.2f}", 'K'])
output_data.append(['Max Deviation Above', f"{max_above:.2f}", 'K'])
output_data.append(['', '', ''])

#comparison of all coatings
output_data.append(['ALL COATINGS COMPARISON', '', ''])
output_data.append(['Coating', 'Alpha', 'Epsilon', 'Optimal Power (W)', 'Energy (kWh)', 'In-Band (%)'])
for name, data in results.items():
    in_band_pct = in_band_fraction[name] * 100
    output_data.append([
        name,
        f"{coatings[name]['alpha']:.4f}",
        f"{coatings[name]['epsilon']:.4f}",
        f"{data['optimal_power_W']:.0f}",
        f"{data['energy_kWh']:.2f}",
        f"{in_band_pct:.2f}",
    ])

#export results to Excel with graphs
print("\n--- Exporting results to Excel ---")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_file = f"satellite_analysis_{timestamp}.xlsx"

#create main analysis sheet
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    #summary data sheet
    output_df = pd.DataFrame(output_data)
    output_df.to_excel(writer, sheet_name='Analysis', index=False, header=False)

#add graphs to excel 
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
wb = load_workbook(excel_file)

#create graphs
fig_eclipse = plot_eclipse_season()
fig_six = create_six_month_plot()

#save figures to bytes
img_eclipse = io.BytesIO()
fig_eclipse.savefig(img_eclipse, format='png', dpi=100)
img_eclipse.seek(0)
plt.close(fig_eclipse)
img_six = io.BytesIO()
fig_six.savefig(img_six, format='png', dpi=100)
img_six.seek(0)
plt.close(fig_six)

#add graph sheets
ws_eclipse = wb.create_sheet('Eclipse Season Graph')
ws_six = wb.create_sheet('Six Month Graph')

#populate sheets with images
xl_img_eclipse = XLImage(img_eclipse)
xl_img_six = XLImage(img_six)

ws_eclipse.add_image(xl_img_eclipse, 'A1')
ws_six.add_image(xl_img_six, 'A1')

#save workbook
wb.save(excel_file)

print(f"Saved complete analysis to: {excel_file}")
print("\n--- Excel Sheets Created ---")
print(f"  1. Analysis - All satellite parameters and coating comparison")
print(f"  2. Eclipse Season Graph - Temperature evolution during eclipse season")
print(f"  3. Six Month Graph - Six-month temperature comparison")

print("\n--- Export Summary ---")
print(f"Satellite dimensions: {height}m H × {width}m W × {depth}m D, Mass: {mass}kg")
print(f"Orbital geometry: Altitude {altitude/1e3:.0f}km, Period {orbital_period/3600:.2f}h")
print(f"Best coating: {best_name}")
print(f"  → Energy: {results[best_name]['energy_kWh']:.2f} kWh/year at {results[best_name]['optimal_power_W']:.0f}W")
print(f"  → Performance: {in_band_fraction[best_name]*100:.2f}% in-band")
print(f"Best balanced: {best_balanced}")
print(f"  → Energy: {results[best_balanced]['energy_kWh']:.2f} kWh/year at {results[best_balanced]['optimal_power_W']:.0f}W")
print(f"  → Performance: {in_band_fraction[best_balanced]*100:.2f}% in-band")
print("\nExport complete!")


