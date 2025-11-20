
from sklearn.linear_model import LinearRegression
import numpy as np
import pandas as pd
from dataclasses import dataclass
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
import matplotlib.pyplot as plt

# solar panel geometry
A_max = 1.0
A_min = 0.0

theta_deg = np.linspace(0, 360, 1000)
time = theta_deg / 15.0  # angle(degrees) to hours

# eclipse during worse day
eclipse_duration_h = 1.18       
eclipse_half_angle = (eclipse_duration_h / 24 * 360) / 2
theta_eclipse_start = 180 - eclipse_half_angle
theta_eclipse_end = 180 + eclipse_half_angle

t_eclipse_start = theta_eclipse_start / 15.0
t_eclipse_end = theta_eclipse_end / 15.0

# sunlight mask
sunlight_mask = np.where(
    (theta_deg >= theta_eclipse_start) & (theta_deg <= theta_eclipse_end),
    0.0,
    1.0
)

# percentage visibility of solar panels function
def area(theta_deg_local):
    theta_rad = np.deg2rad(theta_deg_local)
    A = (A_max - A_min) * np.abs(np.cos(theta_rad)) + A_min
    return A * sunlight_mask


@dataclass
class ModelConfig:
    start_year: int = 2025
    years: int = 15
    eclipse_halfwidth_days: float = 22.5
    out_prefix: str = "solar_eff"
    eclipse_peak_minutes: float = 70.0
    decl_amp_deg: float = 23.44
    decl_phase_shift_day: int = 80
    ecc_amp: float = 0.033
    ecc_perihelion_shift_day: int = 3
    optical_drop_year1: float = 0.07
    cell_drop_year1: float = 0.03
    cell_drop_each_later_year: float = 0.02
    smooth_days: int = 120


# helper functions
def to_doy(dates):

    dates = pd.to_datetime(dates)

    if isinstance(dates, pd.Timestamp):
        return dates.dayofyear
    # DatetimeIndex
    return dates.dayofyear.values

# all efficiency components
def solar_distance_eff(n, amp=0.033, shift=3):
    n = np.asarray(n, dtype=float)
    return 1.0 + amp * np.cos(2 * np.pi * (n - shift) / 365.0)


def solar_angle_eff(n, decl_amp_deg=23.44, phase_shift=80):
    n = np.asarray(n, dtype=float)
    delta_deg = decl_amp_deg * np.sin(2 * np.pi * (n - phase_shift) / 365.0)
    return np.cos(np.deg2rad(delta_deg))


def eclipse_eff(n, centers=(80, 263), halfwidth=22.5, max_minutes=70.0):
    n = np.asarray(n, dtype=float)
    eclipse_minutes = np.zeros_like(n, dtype=float)
    for c in centers:
        d = np.abs(((n - c + 182.5) % 365) - 182.5)
        w = np.where(d <= halfwidth, 0.5 * (1 + np.cos(np.pi * d / halfwidth)), 0.0)
        eclipse_minutes += max_minutes * w
    return 1.0 - eclipse_minutes / (24.0 * 60.0)


def degradation_eff(day_index,
                    optical_drop_year1=0.07,
                    cell_drop_year1=0.03,
                    cell_drop_each_later_year=0.02):
    day_index = np.asarray(day_index, dtype=int)
    year_num = day_index // 365
    optical = np.where(year_num == 0, 1.0, 1.0 - optical_drop_year1)
    cells = (1.0 - cell_drop_year1) * np.power((1.0 - cell_drop_each_later_year), year_num)
    return optical * cells


# combine all to build timeseries
def build_timeseries(cfg: ModelConfig) -> pd.DataFrame:
    start_date = datetime(cfg.start_year, 1, 1)
    total_days = cfg.years * 365

    dates = [start_date + timedelta(days=i) for i in range(total_days)]
    doy = np.array([((to_doy(d) - 1) % 365) + 1 for d in dates])
    day_idx = np.arange(total_days)

    dist = solar_distance_eff(doy, cfg.ecc_amp, cfg.ecc_perihelion_shift_day)
    angle = solar_angle_eff(doy, cfg.decl_amp_deg, cfg.decl_phase_shift_day)
    ecl = eclipse_eff(doy, (80, 263), cfg.eclipse_halfwidth_days, cfg.eclipse_peak_minutes)
    deg = degradation_eff(day_idx, cfg.optical_drop_year1,
                          cfg.cell_drop_year1, cfg.cell_drop_each_later_year)

    total = dist * angle * ecl * deg
    total = np.clip(total, 0.0, 1.2)

    df = pd.DataFrame({
        "date": dates,
        "day_of_year": doy,
        "distance_eff": dist,
        "angle_eff": angle,
        "eclipse_eff": ecl,
        "degradation_eff": deg,
        "total_eff": total
    }).set_index("date")

    return df


def annual_summary(df: pd.DataFrame) -> pd.DataFrame:
    out = df["total_eff"].resample("YE").agg(["min", "mean", "max"])
    out.index = np.arange(1, len(out) + 1)
    out.index.name = "mission_year"
    return out


# plotting functions
def plot_total_with_regressed_degradation(df: pd.DataFrame, cfg: ModelConfig) -> str:
    path = f"{cfg.out_prefix}_total_with_regressed_degradation.png"

    N = len(df)
    # mission-year axis from 1..cfg.years
    t_mission = np.linspace(1.0, float(cfg.years), N)
    # regression on degradation vs years (0..years)
    t_for_reg = np.linspace(0.0, float(cfg.years), N).reshape(-1, 1)

    deg_daily = df["degradation_eff"].values.reshape(-1, 1)
    model = LinearRegression().fit(t_for_reg, deg_daily)
    deg_pred = model.predict(t_for_reg).flatten()

    # build regressed total
    total_regressed = (df["distance_eff"].values * df["angle_eff"].values *
                       df["eclipse_eff"].values * deg_pred)
    total_regressed = np.clip(total_regressed, 0.0, 1.2)

    plt.figure(figsize=(13, 5))
    plt.plot(t_mission, df["total_eff"].values, label="Daily total_eff (raw)",
             alpha=0.7, linewidth=0.8)
    plt.plot(t_mission, total_regressed, label="Total (smoothed via regressed degradation)",
             linewidth=2.2)

    plt.title(f"Total Efficiency — Raw daily vs Smoothed — {cfg.years} years")
    plt.xlabel("Mission Year")
    plt.ylabel("Efficiency (fraction of initial)")
    plt.legend()
    plt.grid(True, linestyle="--", alpha=0.3)

    # annotate regression
    slope = float(np.asarray(model.coef_).ravel()[0])
    intercept = float(np.asarray(model.intercept_).ravel()[0])
    xpos = t_mission[int(max(1, N * 0.02))]
    ypos = max(total_regressed) * 0.98
    plt.text(xpos, ypos, f"deg = {slope:.6f}·t + {intercept:.6f}\n(t in years)",
             fontsize=9, bbox=dict(facecolor="white", alpha=0.8), horizontalalignment='left')

    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()
    return path


def plot_degradation_15yr_with_regression(df: pd.DataFrame, cfg: ModelConfig) -> str:
    path = f"{cfg.out_prefix}_10yr_degradation_regression.png"

    N = len(df)
    t_mission = np.linspace(1.0, float(cfg.years), N)
    t_for_reg = np.linspace(0.0, float(cfg.years), N).reshape(-1, 1)
    deg_daily = df["degradation_eff"].values

    model = LinearRegression().fit(t_for_reg, deg_daily)
    deg_pred = model.predict(t_for_reg)

    slope = float(np.asarray(model.coef_).ravel()[0])
    intercept = float(np.asarray(model.intercept_).ravel()[0])

    plt.figure(figsize=(12, 5))
    plt.plot(t_mission, deg_daily, label="Daily degradation", linewidth=0.9, alpha=0.8)
    plt.plot(t_mission, deg_pred, "k--", linewidth=2.0,
             label=f"Linear fit: deg = {slope:.6f}·t + {intercept:.6f}")

    plt.title("15-Year Degradation (daily) with Linear Regression")
    plt.xlabel("Mission Year")
    plt.ylabel("Degradation efficiency")
    plt.grid(True, linestyle="--", alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()
    return path


# exporting to excel
def export_to_excel(excel_path, annual_df: pd.DataFrame, plot_paths, analysis_write_row=44):

    wb = load_workbook(excel_path)
    # Ensure Analysis sheet exists
    if "Analysis" not in wb.sheetnames:
        raise KeyError("Workbook does not contain 'Analysis' sheet.")
    ws = wb["Analysis"]

    # Write Annual Table at row 48 (safe default)
    start_row = 48
    ws[f"A{start_row}"] = "Mission Year"
    ws[f"B{start_row}"] = "Min"
    ws[f"C{start_row}"] = "Mean"
    ws[f"D{start_row}"] = "Max"

    for i, (yr, row) in enumerate(annual_df.iterrows(), start=start_row + 1):
        ws[f"A{i}"] = int(yr)
        ws[f"B{i}"] = float(row["min"])
        ws[f"C{i}"] = float(row["mean"])
        ws[f"D{i}"] = float(row["max"])

    # Add each plot to its own worksheet so they never overlap
    for idx, p in enumerate(plot_paths, start=1):
        sheet_name = f"Efficiency_Plot_{idx}"
        # replace existing if present
        if sheet_name in wb.sheetnames:
            ws_plot = wb[sheet_name]
        else:
            ws_plot = wb.create_sheet(title=sheet_name)
        img = Image(p)
        # anchor top-left of sheet
        ws_plot.add_image(img, "A1")

    wb.save(excel_path)
    wb.close()


# constants for the solar panel area calculation
internal_energy_use = 10000  # W
I_sun = 1361.0  # W/m^2
absorptivity = 0.86  # default if not present in sheet
eta_panel = 0.40
eta_sys = 0.6037
eta_total = eta_panel * eta_sys

# surface area calculation
def surface_area(energy_heater):
    global energy_J
    energy_J =  (energy_heater) + (internal_energy_use * 86400)  # J needed per day
    power_per_m2 = I_sun * absorptivity * eta_total * area(theta_deg)   

    # integrate over full day
    dt = 86400 / len(power_per_m2)
    energy_per_m2_day = np.sum(power_per_m2) * dt     # J per m² per day
    global area_required 
    area_required = energy_J / energy_per_m2_day

    # 10% margin
    global area_required_10
    area_required_10 = area_required * 1.10
    

    return area_required_10


# excel reading
def load_excel_values(path):
    df = pd.read_excel(path, sheet_name="Analysis", header=None)
    # Try to find 'BEST COATING' label (case-insensitive)
    try:
        idx = df[df[0].astype(str).str.upper() == "BEST COATING"].index[0]
    except Exception:
        idx = None

    if idx is not None:
        # robust extraction with bounds checks
        def safe_get(r, c, default=np.nan):
            try:
                return df.iloc[r, c]
            except Exception:
                return default

        name = safe_get(idx + 1, 1, "")
        #absorptivity = safe_get(idx + 2, 1,absorptivity)
        emissivity = safe_get(idx + 3, 1, np.nan)
        heater_power = safe_get(idx + 4, 1, np.nan)
        annual_kwh = safe_get(idx + 5, 1, np.nan)
        global energy_worst_day
        energy_worst_day = safe_get(idx + 13, 1, np.nan)
    else:
        #absorptivity = absorptivity
        energy_worst_day = np.nan
        name = ""
        emissivity = np.nan
        heater_power = np.nan
        annual_kwh = np.nan

    # ensure floats where needed
    try:
        energy_worst_day = float(energy_worst_day)
    except Exception:
        # if not present, try to compute from annual_kwh if present (kWh/day -> J/day)
        try:
            if not np.isnan(annual_kwh):
                energy_worst_day = float(annual_kwh) * 1000.0  
            else:
                energy_worst_day = np.nan
        except Exception:
            energy_worst_day = np.nan

    return {
        "name": name,
        "absorptivity": absorptivity,
        "emissivity": emissivity,
        "heater_power": heater_power,
        "annual_kwh": annual_kwh,
        "energy_worst_day": energy_worst_day
    }


def main():
    # recursively search current working directory for 'satellite_analysis*.xlsx'
    cwd = os.getcwd()
    excel_pattern = os.path.join(cwd, "**", "satellite_analysis*.xlsx")
    excel_files = glob.glob(excel_pattern, recursive=True)
    if not excel_files:
        raise FileNotFoundError(f"No satellite_analysis*.xlsx files found under {cwd}")
    # pick the most recently modified
    global excel_path
    excel_path = max(excel_files, key=os.path.getctime)

    vals = load_excel_values(excel_path)

    cfg = ModelConfig()
    df = build_timeseries(cfg)

    # Save CSVs
    csv_daily = f"{cfg.out_prefix}_daily_{cfg.start_year}_{cfg.start_year + cfg.years - 1}.csv"
    df.to_csv(csv_daily)

    ann = annual_summary(df)
    ann.to_csv(f"{cfg.out_prefix}_annual_summary.csv")

    # Plots (mission-year x-axis)
    p1 = plot_total_with_regressed_degradation(df, cfg)
    p2 = plot_degradation_15yr_with_regression(df, cfg)

    # Worst-efficiency and timestamp
    worst_eff = float(df["total_eff"].min())
    worst_idx = df["total_eff"].idxmin()  
    worst_calendar_year = int(pd.to_datetime(worst_idx).year)

    # Panel area from energy requirement (we expect energy_worst_day as J/day)
    if np.isnan(vals["energy_worst_day"]):
        raise ValueError("Could not read 'energy_worst_day' from the Excel Analysis sheet. Please ensure the value exists.")
    absorptivity_val = vals.get("absorptivity", absorptivity)
    area_m2 = surface_area(vals["energy_worst_day"])

    usable_power = area_m2 * I_sun * absorptivity_val * eta_total  # W (approx instantaneous at peak)

    # Write results into the Excel Analysis sheet
    wb = load_workbook(excel_path)
    if "Analysis" not in wb.sheetnames:
        raise KeyError("Workbook does not contain 'Analysis' sheet.")
    ws = wb["Analysis"]
    write_row = 44

    ws[f"A{write_row}"] = "Worst Efficiency"
    ws[f"B{write_row}"] = float(worst_eff)

    ws[f"A{write_row + 1}"] = "Worst Year (calendar year)"
    ws[f"B{write_row + 1}"] = int(worst_calendar_year)

    ws[f"A{write_row + 2}"] = "Required Panel Area (m^2) [10% margin included]"
    ws[f"B{write_row + 2}"] = float(area_m2)

    # Save workbook after writing summary facts
    wb.save(excel_path)
    wb.close()

    # Export annual table + plots into same Excel workbook (annual table at row 10 + plots in separate sheets)
    export_to_excel(excel_path, ann, [p1, p2])

    # also return key values
    return {
        "df": df,
        "annual": ann,
        "plots": [p1, p2],
        "excel_path": excel_path,
        "panel_area_m2": area_m2,
        "usable_power_W": usable_power,
        "worst_eff": worst_eff,
        "worst_calendar_year": worst_calendar_year
    }


if __name__ == "__main__":
    out = main()
    # simple confirmation print
    print(f" - Panel area (m^2): {out['panel_area_m2']:.3f}")
    print(f" - Worst eff: {out['worst_eff']:.6f} in year {out['worst_calendar_year']}")


# percentage area calculation where battery is needed
energy_J =  (energy_worst_day) + (internal_energy_use * 86400)
percentage = ((energy_J / 86400) / (I_sun * eta_total * absorptivity * area_required_10))

A_values = area(theta_deg)
mask = A_values < percentage * A_max
indices = np.where(mask)[0]

# root finding intervals
intervals = []
if len(indices) > 0:
    start = indices[0]
    for i in range(1, len(indices)):
        if indices[i] != indices[i - 1] + 1:
            end = indices[i - 1]
            intervals.append((time[start], time[end]))
            start = indices[i]
    intervals.append((time[start], time[indices[-1]]))

print("Intervals when battery is needed:")
for (t1, t2) in intervals:
    print(f"From {t1:.2f} h to {t2:.2f} h")


plt.figure(figsize=(10, 5))
plt.plot(time, area(theta_deg), lw=2, label="Visible Solar Area")

# shade eclipse
plt.axvspan(t_eclipse_start, t_eclipse_end, color='gray', alpha=0.3,
            label="Eclipse (Area=0)")

plt.title('Visible Solar Panel Area vs Time')
plt.xlabel('Time, t (Hours)')
plt.ylabel('Visible surface area, A (%)')
plt.axhline(percentage, color='r', linestyle='--', label="Required Power")
# above the red dotted line the batttery is charging
# below it the battery is being used
plt.xlim(0, 24)
plt.ylim(0, A_max * 1.05)
plt.grid(alpha=0.3)
plt.legend()

# export to excel
area_plot_path = "area_plot.png"
plt.savefig(area_plot_path, dpi=200, bbox_inches='tight')
plt.close()

excel_wb = load_workbook(excel_path)

sheet_name = "Area_Plot"
if sheet_name in excel_wb.sheetnames:
    ws_area = excel_wb[sheet_name]
else:
    ws_area = excel_wb.create_sheet(sheet_name)

img = Image(area_plot_path)
ws_area.add_image(img, "A1")

excel_wb.save(excel_path)
excel_wb.close()

print("Export complet!")

os.remove("area_plot.png")
os.remove("solar_eff_daily_2025_2039.csv")
os.remove("solar_eff_total_with_regressed_degradation.png")
os.remove("solar_eff_annual_summary.csv")
os.remove("solar_eff_10yr_degradation_regression.png")

